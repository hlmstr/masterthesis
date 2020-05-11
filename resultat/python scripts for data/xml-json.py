import xmltodict, json
from openpyxl import Workbook
import numpy as np
test = [21897,
      82225,
      109777,
      123009,
      126892,
      133167,
      147210,
      165139,
      170871,
      210040,
      221431,
      94541,
      137158,
      154951,
      202739,
      164678,
      206541,
      200082,
      186314,
      154773,
      180199,
      160023,
      172105,
      155717,
      138954,
      141777,
      111994,
240160,
240085,
240123]
print(np.percentile(test,99))
microResults = []
monoResults = []
documents = ["micro_1-8000.xml","micro_10000-12000.xml","mono_1-8000.xml","mono_10000-12000.xml"]


for doc in documents:
    print("start loading doc: " + doc)

    with open(doc) as xml_file:
        my_dict=xmltodict.parse(xml_file.read())
    xml_file.close()
    print("loading done")

    counter = 0


    for requestResult in my_dict['testResults']['httpSample']:
        result = {}
        # print('latency: ' + str(requestResult['@lt']))
        # print('request group: ' + str(requestResult['@lb']))
        if("Gateway" not in str(requestResult['@rm'])):
            if("micro" in str(requestResult['@lb'])):
                result["latency"] = int(requestResult['@lt']) + int(requestResult['@ct'])
                result["thread group"] = requestResult["@lb"]
                result["active threads"] = int(requestResult["@na"])
                jointime = int(json.loads(requestResult['responseData']['#text'])['jointime'])
                result["jointime"] = jointime
                microResults.append(result)
            else:
                result["latency"] = int(requestResult['@lt']) + int(requestResult['@ct'])
                result["thread group"] = requestResult["@lb"]
                result["active threads"] = int(requestResult["@na"])
                result["jointime"] = "null"
                monoResults.append(result)
    # print(json.dumps(monoResults, indent = 2))
    # print("-----")
    # print(json.dumps(microResults, indent = 2))

        # print(json.loads(my_dict['testResults']['httpSample'][1]['responseData']['#text'])['jointime'])
        # print(my_dict['testResults']['httpSample'][1]['@lt'])


# print(json.dumps(monoResults, indent = 2))
print("micro" +str(len(microResults)))
print("mono" + str(len(monoResults)))



results = []
results.append(monoResults)
results.append(microResults)
monoAggregation = {}
microAggregation = {}
mono = True
c1=0
c2=0
data = {}

for dataset in results:
    if(mono):
        for response in dataset:
            if(response["thread group"] not in data.keys()): 
                data[response["thread group"]] = {"thread group": response["thread group"], "latency": int(response["latency"]),"counter": 1, "response_times": [int(response["latency"])]}
            else:
                data[response["thread group"]]["latency"] = data[response["thread group"]]["latency"] + int(response["latency"])
                data[response["thread group"]]["counter"] = data[response["thread group"]]["counter"] + 1
                data[response["thread group"]]["response_times"].append(int(response["latency"]))
                data[response["thread group"]]["averageresponse"] = int(data[response["thread group"]]["latency"])/int(data[response["thread group"]]["counter"])
    else:
        for response in dataset:
            response["latency"]
            if(response["thread group"] not in data.keys()): 
                data[response["thread group"]] = {"thread group": response["thread group"], "latency": int(response["latency"]),"jointime": response["jointime"],"counter": 1, "response_times" : [int(response["latency"])]}
            else:
                data[response["thread group"]]["latency"] = data[response["thread group"]]["latency"] + int(response["latency"])
                data[response["thread group"]]["counter"] = data[response["thread group"]]["counter"] + 1
                data[response["thread group"]]["response_times"].append(int(response["latency"]))
                data[response["thread group"]]["averagejoin"] = int(data[response["thread group"]]["jointime"])/int(data[response["thread group"]]["counter"])

                if("null" not in str(response["jointime"])):
                    data[response["thread group"]]["jointime"] = data[response["thread group"]]["jointime"] + response["jointime"] 
                else:
                    data["jointime"] = "null"
                data[response["thread group"]]["averageresponse"] = int(data[response["thread group"]]["latency"])/int(data[response["thread group"]]["counter"])
    mono = False

print(data.keys())
for key in data:
    print(data[key]) 
    data[key]["90p"] = np.percentile(data[key]["response_times"], 90)
    data[key]["95p"] = np.percentile(data[key]["response_times"], 95)
    data[key]["99p"] = np.percentile(data[key]["response_times"], 99)


print(json.dumps(data, indent = 2))







wb = Workbook()
wb.create_sheet("monolith")
wb.create_sheet("microservices")
raw_data_sheets = ["monolith", "microservices"]
for i in range(len(results)):
    result = results[i]
    current_worksheet = wb.get_sheet_by_name(raw_data_sheets[i])
    curr_col = 1
    curr_row = 1
    current_worksheet.cell(row=curr_row, column=curr_col).value = "monolith"
    current_worksheet.cell(row=curr_row, column=curr_col+1).value = "latency"
    current_worksheet.cell(row=curr_row, column=curr_col+2).value = "thread group"
    current_worksheet.cell(row=curr_row, column=curr_col+3).value = "active threads"
    current_worksheet.cell(row=curr_row, column=curr_col+4).value = "join time"

    for response in result:
        curr_row = curr_row + 1
        current_worksheet.cell(row=curr_row, column=curr_col+1).value = response["latency"]
        current_worksheet.cell(row=curr_row, column=curr_col+2).value = response["thread group"]
        current_worksheet.cell(row=curr_row, column=curr_col+3).value = response["active threads"]
        current_worksheet.cell(row=curr_row, column=curr_col+4).value = response["jointime"]

wb.create_sheet("data")
current_worksheet = wb.get_sheet_by_name("data")
curr_col = 1
curr_row = 1
current_worksheet.cell(row=curr_row, column=curr_col).value = "threadgroup"
current_worksheet.cell(row=curr_row, column=curr_col+1).value = "counter"
current_worksheet.cell(row=curr_row, column=curr_col+2).value = "p90"
current_worksheet.cell(row=curr_row, column=curr_col+3).value = "p95"
current_worksheet.cell(row=curr_row, column=curr_col+4).value = "p99"
current_worksheet.cell(row=curr_row, column=curr_col+5).value = "responsetimeaverage"
current_worksheet.cell(row=curr_row, column=curr_col+6).value = "jointimeaverage"




for key in data:
    curr_row= curr_row + 1
    current_worksheet.cell(row=curr_row, column=curr_col).value = key
    current_worksheet.cell(row=curr_row, column=curr_col+1).value = data[key]["counter"]
    current_worksheet.cell(row=curr_row, column=curr_col+2).value = data[key]["90p"]
    current_worksheet.cell(row=curr_row, column=curr_col+3).value = data[key]["95p"]
    current_worksheet.cell(row=curr_row, column=curr_col+4).value = data[key]["99p"]
    current_worksheet.cell(row=curr_row, column=curr_col+5).value = data[key]["averageresponse"]
    if("micro" in key):
        current_worksheet.cell(row=curr_row, column=curr_col+6).value = data[key]["averagejoin"]
 

wb.save("results.xlsx")

print(wb.sheetnames)

# print(json_data)