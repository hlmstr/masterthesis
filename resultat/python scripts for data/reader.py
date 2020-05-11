from openpyxl import load_workbook
import json
print("Loading doc")
wb = load_workbook(filename="EX_2_TP_calc_no_fails.xlsx")
print("Loading doc done")

print(wb.sheetnames)
allowed_thread_count = [100,150,200,250,300,350,400,450,500]
current_column = 1
for sheet in wb.sheetnames:
    if(sheet=="results"):
        continue
    current_worksheet = wb.get_sheet_by_name(sheet)
    data = {}
    max=current_worksheet.max_row
    for row in current_worksheet.iter_rows(min_row=0, max_row=max, max_col=2, values_only=True):
        if(row[0] in allowed_thread_count):
            # print("row in allowed",row[0])
            if(row[0] not in data.keys()):
                data[row[0]] = {'response_time': row[1], 'count':1}
                # print("adding new", tp_mono_data[row[0]])

            else:
                data[row[0]]['count'] = data[row[0]]['count']+1
                data[row[0]]['response_time'] = data[row[0]]['response_time'] + row[1]
                data[row[0]]['average'] = data[row[0]]['response_time'] / data[row[0]]['count']

                # print("Existed, count",  tp_mono_data[row[0]]['count'])
                # print("Existed responsetime", tp_mono_data[row[0]]['response_time']
    print(json.dumps(data, indent=3))

    if "results" not in wb.sheetnames:
        wb.create_sheet("results")
    current_worksheet = wb.get_sheet_by_name("results")

    current_row = 1
    
    architecture_column = current_worksheet.cell(row=current_row, column=current_column)

    thread_count_column = current_worksheet.cell(row=current_row, column=current_column+1)
    average_column = current_worksheet.cell(row=current_row, column=current_column+2)
    count_column = current_worksheet.cell(row=current_row, column=current_column+3)
    tp_column = current_worksheet.cell(row=current_row, column=current_column+4)

    architecture_column.value = sheet
    thread_count_column.value ='thread_count'
    average_column.value ='average'
    count_column.value ='count'
    tp_column.value ='tp'


    for thread_count in data:
        current_row = current_row+1

        average = data[thread_count]['average']
        count = data[thread_count]['count']

        thread_count_column = current_worksheet.cell(row=current_row, column=current_column+1)
        average_column = current_worksheet.cell(row=current_row, column=current_column+2)
        count_column = current_worksheet.cell(row=current_row, column=current_column+3)
        tp_column = current_worksheet.cell(row=current_row, column=current_column+4)

        thread_count_column.value = thread_count
        average_column.value = average
        count_column.value = count
        tp_column.value= thread_count*(1000/average)
    current_column = current_column + 6 
print("start to save")
wb.save("EX_2_tp_calc_results.xlsx")
print("save done")












